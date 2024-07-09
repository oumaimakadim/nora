from tkinter import *
from tkinter import ttk 
import tkinter as tk
from tkinter import messagebox 
import openpyxl as xl
import app
import customtkinter
import xlrd

appe = Tk()
appe.title('Employee Management System')
appe.geometry('1500x620')
appe.config(bg='#161C25')
appe.resizable(False,False)
font1 = ('Arial', 20, 'bold')


id_label=customtkinter.CTkLabel(appe, text='Id',text_color='#fff', bg_color ='#161C25')
id_label.place(x=20,y=20)
id_entry=customtkinter.CTkEntry(appe, fg_color ='#fff',bg_color ='#161C25',width=180)
id_entry.place(x=180,y=20)
name_label=customtkinter.CTkLabel(appe, text='Name', text_color='#fff', bg_color ='#161C25')
name_label.place(x=20,y=80)
name_entry=customtkinter.CTkEntry(appe, bg_color ='#161C25', width=180)
name_entry.place(x=180,y=80)

satisf_label=customtkinter.CTkLabel(appe, text='satisfaction', text_color='#fff', bg_color ='#161C25')
satisf_label.place(x=20,y=140)



options= ['satisfied','yes, but i have a problem']
Variable1=StringVar
satisf_options= customtkinter.CTkComboBox(appe,text_color='#000', fg_color ='#fff', dropdown_hover_color='#0C9295',button_color='#0C9295',button_hover_color='#0C9295',width=180,variable=Variable1,values=options ,state='readonly')
satisf_options.set('satisfied')
satisf_options.place(x=180, y=140)

problem_label=customtkinter.CTkLabel(appe, text='problem', text_color='#fff', bg_color ='#161C25')
problem_label.place(x=20,y=200)
problem_entry=customtkinter.CTkEntry(appe, fg_color ='#fff', bg_color ='#161C25',width=180)
problem_entry.place(x=180,y=200)
compo_label=customtkinter.CTkLabel(appe, text='Component attacked', text_color='#fff', bg_color ='#161C25')
compo_label.place(x=20,y=260)
compo_entry=customtkinter.CTkEntry(appe, fg_color ='#fff', bg_color ='#161C25',width=180)
compo_entry.place(x=180,y=260)

def submit():
   if id_entry.get() !='' and name_entry.get() !='' and name_entry.get()!='' and satisf_options.get()!='' and problem_entry.get()!='' and compo_entry.get()!=0:
       file = xl.load_workbook("C://Users/Nom/Desktop/data.xlsx") 
       sheet = file['Sheet1']
       id_value = id_entry.get() 
       name_value = name_entry.get()
       satisf_value = satisf_options.get()
       problem_value = problem_entry.get()
       compo_value = compo_entry.get()

       sheet.cell(column=1, row=sheet.max_row+1, value=id_value)
       sheet.cell (column=2, row=sheet.max_row,value=name_value) 
       sheet.cell (column=3, row=sheet.max_row,value=satisf_value) 
       sheet.cell(column=4,row=sheet.max_row,value=problem_value)
       sheet.cell(column=5, row=sheet.max_row,value=compo_value)
       file.save("C://Users/Nom/Desktop/data.xlsx")
       messagebox.showinfo('Success', 'Data has been saved.')
   else:
       messagebox.showerror('Error', 'Enter all the data.')

add_button = customtkinter.CTkButton(appe, command=submit, text_color='#000', fg_color ='#05A312', text='Add Problem', hover_color= '#00850B', bg_color='#161C25', corner_radius=15, width=260)
add_button.place(x=20,y=310)
clear_button = customtkinter.CTkButton(appe, text_color='#000', fg_color ='#FFFF00', text='New Problem', hover_color= '#00850B', bg_color='#161C25', corner_radius=15, width=260)
clear_button.place(x=20,y=360)
up_button = customtkinter.CTkButton(appe, text_color='#000', fg_color ='#DFAF2C', text='Update Problem', hover_color= '#00850B', bg_color='#161C25', corner_radius=15, width=260)
up_button.place(x=300,y=360)
dele_button = customtkinter.CTkButton(appe, text_color='#000', fg_color ='#f00020', text='Delete Problem', hover_color= '#00850B', bg_color='#161C25', corner_radius=15, width=260)
dele_button.place(x=580,y=360)
      

style = ttk.Style(appe)

style.theme_use('clam')
style.configure('treeview',foreground= '#fff',background= '#000',fieldbackground='#313837')
style.map('treeview',background=[('selected', '#1A8F2D')])

tree = ttk.Treeview(appe,height=15)



tree['columns'] = ('ID', 'Name', 'Satisfaction', 'Problem', 'Component attacked')
tree.column('#0', width=0, stretch=tk.NO) 
tree.column('ID', anchor=tk.CENTER, width=120)
tree.column('Name', anchor=tk.CENTER, width=120)
tree.column('Satisfaction', anchor=tk.CENTER, width=120)
tree.column('Problem', anchor=tk.CENTER, width=100)
tree.column('Component attacked', anchor=tk.CENTER, width=120)

tree.heading('ID', text='ID')
tree.heading('Name', text='Name')
tree.heading('Satisfaction', text='Satisfaction')
tree.heading('Problem', text='Problem')
tree.heading('Component attacked', text='Component attacked')

tree.place(x=900,y=20)

appe.mainloop()
